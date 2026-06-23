#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Descarga las tablas oficiales de crecimiento de la OMS (Child Growth Standards
2006, 0–5 años; Growth Reference 2007, 5–19 años) y genera los JSON en formato
LMS que consume el backend (src/data/oms/<indicador>.json).

Indicadores generados: peso_edad (P/E), talla_edad (T/E), imc_edad (IMC/E),
pce_edad (PCe/E). El perímetro/peso-talla (P/T) NO se incluye aquí porque la OMS
lo publica como dos estándares distintos (longitud 0–2 y talla 2–5); ver README.

NO inventa datos: solo descarga y transforma la fuente oficial. Incluye una
autovalidación: comprueba que el Z calculado en los puntos SD−2 / SD0 / SD+2 de
la propia OMS dé ≈ −2 / 0 / +2.

Uso:  python3 scripts/descargarTablasOMS.py
"""

import json
import math
import os
import subprocess
import sys
import tempfile

import openpyxl

DIAS_POR_MES = 30.4375
OUT_DIR = os.path.join(os.path.dirname(__file__), "..", "src", "data", "oms")

B6 = "https://cdn.who.int/media/docs/default-source/child-growth/child-growth-standards/indicators"
B7 = "https://cdn.who.int/media/docs/default-source/child-growth/growth-reference-5-19-years"

# indicador -> lista de (url, sexo, unidad_x)   unidad_x: "day" | "month"
FUENTES = {
    "imc_edad": [
        (f"{B6}/body-mass-index-for-age/expanded-tables/bfa-boys-zscore-expanded-tables.xlsx", "M", "day"),
        (f"{B6}/body-mass-index-for-age/expanded-tables/bfa-girls-zscore-expanded-tables.xlsx", "F", "day"),
        (f"{B7}/bmi-for-age-(5-19-years)/bmi-boys-z-who-2007-exp.xlsx", "M", "month"),
        (f"{B7}/bmi-for-age-(5-19-years)/bmi-girls-z-who-2007-exp.xlsx", "F", "month"),
    ],
    "talla_edad": [
        (f"{B6}/length-height-for-age/expandable-tables/lhfa-boys-zscore-expanded-tables.xlsx", "M", "day"),
        (f"{B6}/length-height-for-age/expandable-tables/lhfa-girls-zscore-expanded-tables.xlsx", "F", "day"),
        (f"{B7}/height-for-age-(5-19-years)/hfa-boys-z-who-2007-exp.xlsx", "M", "month"),
        (f"{B7}/height-for-age-(5-19-years)/hfa-girls-z-who-2007-exp.xlsx", "F", "month"),
    ],
    "peso_edad": [
        (f"{B6}/weight-for-age/expanded-tables/wfa-boys-zscore-expanded-tables.xlsx", "M", "day"),
        (f"{B6}/weight-for-age/expanded-tables/wfa-girls-zscore-expanded-tables.xlsx", "F", "day"),
    ],
    "pce_edad": [
        (f"{B6}/head-circumference-for-age/expanded-tables/hcfa-boys-zscore-expanded-tables.xlsx", "M", "day"),
        (f"{B6}/head-circumference-for-age/expanded-tables/hcfa-girls-zscore-expanded-tables.xlsx", "F", "day"),
    ],
}

REFERENCIA = {
    "imc_edad": "OMS 2006 (0-5a) / OMS 2007 (5-19a)",
    "talla_edad": "OMS 2006 (0-5a) / OMS 2007 (5-19a)",
    "peso_edad": "OMS 2006 (0-5a)",
    "pce_edad": "OMS 2007 (0-5a)",
}


def z_lms(x, L, M, S):
    return math.log(x / M) / S if L == 0 else (pow(x / M, L) - 1) / (L * S)


def descargar_xlsx(url):
    """Descarga con curl (usa los certificados del sistema) a un archivo temporal."""
    fd, tmp = tempfile.mkstemp(suffix=".xlsx")
    os.close(fd)
    subprocess.run(["curl", "-sL", "--fail", "--max-time", "60", "-o", tmp, url], check=True)
    return tmp


def leer_tabla(buf, unidad_x):
    """Devuelve (filas, autocheck) donde filas=[{x,L,M,S}] con x en meses (o cm),
    y autocheck = lista de (x, sd2neg, M, sd2pos) para validar el Z."""
    wb = openpyxl.load_workbook(buf, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    it = ws.iter_rows(values_only=True)
    header = [str(c).strip().lower() if c is not None else "" for c in next(it)]

    def idx(*names):
        for n in names:
            if n in header:
                return header.index(n)
        return None

    xi = idx("month", "day", "age", "length", "height")
    li, mi, si = header.index("l"), header.index("m"), header.index("s")
    sd2n, sd0, sd2p = idx("sd2neg"), idx("sd0"), idx("sd2")

    filas, checks, vistos = [], [], set()
    for row in it:
        try:
            xraw = float(row[xi]); L = float(row[li]); M = float(row[mi]); S = float(row[si])
        except (TypeError, ValueError):
            continue
        # normalizar X a meses si viene en días; conservar enteros (1 fila/mes)
        if unidad_x == "day":
            mes = round(xraw / DIAS_POR_MES)
            x = mes
        else:
            x = round(xraw, 4)
        if x in vistos:
            continue
        vistos.add(x)
        filas.append({"x": x, "L": round(L, 6), "M": round(M, 6), "S": round(S, 6)})
        if sd2n is not None and sd0 is not None and sd2p is not None and len(checks) < 3:
            try:
                checks.append((x, float(row[sd2n]), float(row[sd0]), float(row[sd2p]), L, M, S))
            except (TypeError, ValueError):
                pass
    return filas, checks


def main():
    os.makedirs(OUT_DIR, exist_ok=True)
    total_check_ok = 0
    total_check = 0

    for indicador, fuentes in FUENTES.items():
        datos = {"M": [], "F": []}
        print(f"\n[{indicador}]")
        for url, sexo, unidad in fuentes:
            try:
                buf = descargar_xlsx(url)
                filas, checks = leer_tabla(buf, unidad)
                os.remove(buf)
            except Exception as e:
                print(f"  ✗ {os.path.basename(url)}: {e}")
                continue
            datos[sexo].extend(filas)
            print(f"  ✓ {os.path.basename(url)} [{sexo}] {len(filas)} filas")
            # autovalidación: Z en SD-2/SD0/SD+2 debe dar ≈ -2/0/+2
            for (x, v2n, v0, v2p, L, M, S) in checks:
                for val, esperado in ((v2n, -2), (v0, 0), (v2p, 2)):
                    total_check += 1
                    z = z_lms(val, L, M, S)
                    if abs(z - esperado) <= 0.02:
                        total_check_ok += 1
                    else:
                        print(f"    ⚠ autovalidación x={x}: Z({val})={z:.3f} esperado {esperado}")

        for s in ("M", "F"):
            datos[s].sort(key=lambda r: r["x"])
            # dedup por x (por si 2006 y 2007 coinciden en el borde)
            vistos, limpio = set(), []
            for r in datos[s]:
                if r["x"] in vistos:
                    continue
                vistos.add(r["x"]); limpio.append(r)
            datos[s] = limpio

        salida = {
            "indicador": indicador,
            "referencia": REFERENCIA[indicador],
            "variableX": "edad_meses",
            "uso_clinico": True,
            "fuente": "Descargado de cdn.who.int con scripts/descargarTablasOMS.py",
            "datos": datos,
        }
        ruta = os.path.join(OUT_DIR, f"{indicador}.json")
        with open(ruta, "w", encoding="utf-8") as f:
            json.dump(salida, f, ensure_ascii=False, indent=1)
        print(f"  → {os.path.relpath(ruta)} (M:{len(datos['M'])} F:{len(datos['F'])})")

    print(f"\nAutovalidación OMS (Z en puntos SD): {total_check_ok}/{total_check} correctos")
    if total_check and total_check_ok == total_check:
        print("✅ Tablas OMS descargadas y verificadas contra sus propios puntos SD.")
    else:
        print("⚠ Revisar discrepancias arriba.")


if __name__ == "__main__":
    sys.exit(main())
